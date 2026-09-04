# -*- coding: utf-8 -*-
"""Draft entity/relationship model for Dona Dom — Building and Unit.
3 Sep 2026."""

import os

from openpyxl import Workbook
from openpyxl.styles import Font as _Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def Font(**kw):
    """Every cell in this workbook is Arial unless told otherwise."""
    kw.setdefault("name", "Arial")
    return _Font(**kw)


INK        = "1B2A38"
HEAD_BG    = "27405A"
GROUP_BG   = "DCE4EC"
NOTE_BG    = "F3F1E9"
RULE       = "B9C4CE"
GOOD       = "2F6B45"
WARN       = "9A5B14"
STOP       = "A33A1E"

thin = Side(style="thin", color=RULE)
BORDER = Border(bottom=thin)

wb = Workbook()


def sheet(title, headers, widths, rows, freeze="A2"):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 2
    for row in rows:
        if isinstance(row, tuple) and row and row[0] == "__GROUP__":
            c = ws.cell(row=r, column=1, value=row[1])
            c.font = Font(bold=True, size=10, color=INK)
            for i in range(1, len(headers) + 1):
                ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor=GROUP_BG)
            ws.row_dimensions[r].height = 20
            r += 1
            continue
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(size=10, color=INK)
            c.border = BORDER
        r += 1

    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{r-1}"
    return ws


# ---------------------------------------------------------------- READ ME
ws = wb.active
ws.title = "READ ME"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 112

lines = [
    ("h1", "Dona Dom — Building and Unit"),
    ("sub", "3 September 2026. Minimal fields on purpose — the goal of this pass is to get the "
            "relationships right, then add fields on top of a shape that already holds. Nothing "
            "here is published; this workbook is handed over as a file."),
    ("gap", ""),
    ("h2", "Numbering convention"),
    ("n1", "Relationship numbers are append-only. A new relationship is appended (R15, R16, …) and "
           "never inserted, so R1–R14 keep the numbers cited elsewhere — including in the Hebrew "
           "workbook. Entity numbers (E1–E14) carry no such guarantee: nothing cross-references them."),
    ("gap", ""),
    ("h2", "The one idea that keeps this simple"),
    ("p",  "A building is not a bag of apartments. It is a set of SPACES — some private and leasable "
           "(apartments), some shared (lobby, stairwell, roof), some technical (pump room, "
           "electrical room), some outside (garden, gate), and some assignable (parking bays, "
           "storage rooms). Everything that can break, be inspected or be handed over is an ASSET, "
           "and every asset sits in exactly one space."),
    ("p",  "That means common areas, safety equipment and operational utilities do NOT become three "
           "new families of tables. Common areas are spaces. Fire extinguishers, sprinklers, pumps, "
           "elevators and boilers are assets that happen to sit in a shared or technical space. One "
           "table for places, one for things. Every query about a building is then the same query."),
    ("gap", ""),
    ("h2", "Why that shape pays for itself immediately"),
    ("p",  "Responsibility falls out of it. An asset in a UNIT space can be tenant, operator or "
           "contractor. An asset in a COMMON, TECHNICAL or EXTERIOR space is never the tenant's — "
           "the rule reads the space kind and stops. We get half the responsibility matrix from the "
           "location of the thing, without writing a rule per asset type."),
    ("gap", ""),
    ("h2", "Five rules the draft obeys everywhere"),
    ("n1", "1.  The current tenant is a VIEW, never a column. A unit does not hold a tenant. A tenancy "
           "holds a unit and a date range, and 'who lives there' is resolved as today ∈ [start, end]. "
           "This is what makes tenant isolation safe — see the Relationships sheet, row R6."),
    ("n1", "2.  Nothing is overwritten. A new lease does not replace the old one, it follows it. A "
           "corrected rule supersedes by date. Last year's decision can still be reconstructed."),
    ("n1", "3.  No money. Not one amount, anywhere in this model. Rent, arrears and invoices stay in "
           "Priority. We record whether an obligation is SATISFIED, never what it costs."),
    ("n1", "4.  Every fact points at the paper it came from. Documents are linked, not summarised."),
    ("n1", "5.  Structure carries meaning, so text does not have to. Where a field is a short fixed "
           "list, the list is in the draft. Free text is the exception."),
    ("gap", ""),
    ("h2", "What is deliberately NOT in this pass"),
    ("p",  "Service calls, visits, providers, the state machine and the agent. You asked for building "
           "then unit, and those sit downstream of both. Provider appears once, as a stub, only because "
           "an asset under warranty has to point at whoever is on the hook for it during תקופת הבדק. "
           "Nothing else about the service loop is touched."),
    ("p",  "Also absent: rent, deposits, arrears, indexation, and anything an accountant would "
           "recognise. That is Priority's, permanently."),
    ("gap", ""),
    ("h2", "How to read the sheets"),
    ("p",  "ENTITIES — everything that exists, and what one row of each actually is. Read this first; "
           "if a grain is wrong, everything downstream is wrong."),
    ("p",  "RELATIONSHIPS — the heart of it. Every line between two entities, in plain words, with what "
           "breaks if we get it backwards."),
    ("p",  "FIELDS — the minimum each entity carries. Grouped by entity, filterable. This is the sheet "
           "that grows: add rows, don't reshape."),
    ("p",  "ADMIN VIEWS — what an admin actually sees on a building screen, a unit screen and the "
           "settings screen, and which parts are stored versus derived. The test of the model is "
           "whether these screens fall out of it without special cases."),
    ("p",  "DECISIONS — the six shaping decisions, each stated as the rule it creates and why that "
           "rule holds. Kept in the workbook so they are read alongside the tables they explain."),
]

r = 2
for kind, text in lines:
    c = ws.cell(row=r, column=2, value=text)
    if kind == "h1":
        c.font = Font(bold=True, size=18, color=INK)
        ws.row_dimensions[r].height = 26
    elif kind == "sub":
        c.font = Font(size=10, italic=True, color="4C5964")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 42
    elif kind == "h2":
        c.font = Font(bold=True, size=12, color=HEAD_BG)
        ws.row_dimensions[r].height = 24
    elif kind == "gap":
        ws.row_dimensions[r].height = 8
    else:
        c.font = Font(size=10, color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 13 * (len(text) // 100 + 1))
    r += 1


# ---------------------------------------------------------------- ENTITIES
entities = [
    ("E1", "Project", "One דירה להשכיר tender or development, above the buildings in it.",
     "OPTIONAL container. Groups buildings that share a tender, a garden, a gate or a set of terms. "
     "A standalone building does not need one.",
     "NEW — D2, optional"),
    ("E2", "Building", "One physical building with one entrance core and one address.",
     "The container. Holds spaces, and carries the handover date that starts תקופת הבדק.",
     "New in detail"),
    ("E3", "Space", "One addressable place inside a building — an apartment, the lobby, the roof, "
     "a pump room, a parking bay, a storage room.",
     "The single answer to 'where is this?'. Replaces the need for separate common-area, safety and "
     "utility structures. Every asset and every service call points at a space.",
     "SETTLED — D1, in"),
    ("E4", "Unit", "One leasable apartment. A 1-to-1 extension of the Space that is that apartment.",
     "Everything that is true of an apartment but not of a stairwell: door number, rooms, area, "
     "ממ\"ד, and the fact that it can be leased.",
     "Approved, extended"),
    ("E5", "Party", "One person or one company we have any dealings with.",
     "Tenants, co-tenants, guarantors, contacts. One party, many roles over time — a party is never "
     "'a tenant', it plays the tenant role in a tenancy.",
     "Approved"),
    ("E6", "PartyContact", "One channel (phone or email) belonging to one party, over one period.",
     "The agent's front door. A WhatsApp message arrives as a phone number and must resolve to exactly "
     "one party. Dated, because numbers get recycled.",
     "NEW — split out"),
    ("E7", "Tenancy", "One lease term on one unit, with a start and an end.",
     "The join between a unit and the people in it, in time. This is the 'tenancy record' — it lives "
     "here, not on the unit.",
     "Approved"),
    ("E8", "TenancyParty", "One person's role on one tenancy.",
     "A lease usually has more than one name on it. Distinguishes primary tenant from co-tenant, "
     "guarantor (ערב) and non-signing occupant — and controls who the agent may talk to.",
     "NEW — split out"),
    ("E9", "Obligation", "One recurring duty attached to one tenancy, with a validity period.",
     "The 'liabilities': ארנונה, contents insurance, bank guarantee, utility accounts, ועד בית. "
     "Status only — never an amount.",
     "NEW"),
    ("E10", "ObligationType", "One kind of obligation that can exist, managed by an admin on a screen.",
     "So the list can grow without a release. Seeded with the five above; ועד בית variants and "
     "sub-metered water get added here, by an admin, not by a developer.",
     "NEW — D5"),
    ("E11", "Asset", "One thing that can break, be inspected, or be handed over — in exactly one space.",
     "Air conditioner, water heater, elevator, pump, fire extinguisher, sprinkler riser, gate motor, "
     "intercom. Carries the warranty date that drives the ternary responsibility decision.",
     "Approved, extended"),
    ("E12", "Document", "One file, copied and hashed at ingest.",
     "Lease, amendment, termination notice, ארנונה bill, insurance certificate, ID, bank guarantee, "
     "handover protocol, inspection certificate. The evidence behind every fact above.",
     "Approved"),
    ("E13", "DocumentLink", "One binding between one document and one entity.",
     "A signed lease is evidence about a tenancy AND its unit AND two parties. One document, several "
     "bindings — so the binding is its own row.",
     "NEW — split out"),
    ("E14", "Provider", "A contractor, supplier or in-house crew. STUB ONLY in this pass.",
     "Present only so an asset under warranty can point at who must fix it during תקופת הבדק. "
     "Everything else about providers is out of scope here.",
     "Out of scope"),
]
sheet("ENTITIES",
      ["#", "Entity", "One row = ", "What it is for", "Status"],
      [5, 16, 40, 62, 17],
      [(a, b, c, d, e) for a, b, c, d, e in entities])


# ---------------------------------------------------------------- RELATIONSHIPS
rels = [
    ("R1", "Building", "has many", "Space", "1 : N", "Space.building_id → Building.building_id",
     "Every space belongs to exactly one building. A building with no spaces is a building we have not "
     "surveyed yet.",
     "Nothing much. This one is safe."),
    ("R2", "Space", "is, when leasable", "Unit", "1 : 0..1", "Unit.unit_id = Space.space_id (shared key)",
     "An apartment is a space that can be leased. The apartment-only facts live in Unit; the "
     "where-is-it facts live in Space. A lobby has a Space row and no Unit row.",
     "If we flatten these into one table, every query has to remember that half the rows are not "
     "apartments. If we separate them completely, an asset needs two nullable location columns. "
     "This shape avoids both."),
    ("R3", "Space", "holds many", "Asset", "1 : N", "Asset.space_id → Space.space_id",
     "THE KEY LINE. A fire extinguisher in the stairwell, an air conditioner in apartment 12 and the "
     "main water pump in the pump room are all the same relationship. One join answers 'what is in "
     "this building', 'what is in this apartment', 'what is due for inspection'.",
     "This is what stops common areas, safety equipment and utilities from becoming three parallel "
     "systems that each need their own screen, their own service call type and their own report."),
    ("R4", "Space", "kind constrains", "Asset responsibility", "rule, not a table", "reads Space.space_kind",
     "UNIT is the only kind that can ever be the tenant's. COMMON, TECHNICAL and EXTERIOR never are. "
     "PARKING and STORAGE follow the unit they are assigned to; unassigned, they are treated as "
     "COMMON. That is the whole rule.",
     "Getting this wrong means billing a tenant for the lobby light — or leaving a bay nobody owns. "
     "The rule is four lines because the location already carries the meaning."),
    ("R5", "Unit", "has many over time", "Tenancy", "1 : N", "Tenancy.unit_id → Unit.unit_id",
     "A unit accumulates tenancies. The old one is never deleted or overwritten when a new tenant "
     "moves in — it ends, and the next one starts.",
     "Overwriting destroys the history a dispute needs. 'Who lived here in March 2025' must always be "
     "answerable."),
    ("R6", "Tenancy", "resolves", "current occupancy", "VIEW, never a column", "today ∈ [start_date, end_date]",
     "There is no current_tenant field on Unit, and there must never be one. Who lives in a unit today "
     "is computed from dates, every time it is asked.",
     "This is the isolation constraint. A stored pointer goes stale silently — someone moves out, the "
     "column still says their name, and the agent answers the wrong person about their own home. "
     "A date range cannot go stale."),
    ("R7", "Tenancy", "involves", "Party", "N : M, via TenancyParty", "TenancyParty(tenancy_id, party_id)",
     "Two spouses on one lease are two parties on one tenancy. One person renting three units over "
     "five years is one party on three tenancies.",
     "Without the join table, a second signatory has nowhere to go and gets typed into a notes field, "
     "where no rule can read them."),
    ("R8", "TenancyParty", "carries", "role and reachability", "attribute + hard rule", "role, is_service_contact",
     "Primary tenant, co-tenant, guarantor (ערב), occupant. LOCKED BY D4: a GUARANTOR is never a "
     "service contact. The flag is forced false for that role and cannot be set true — not by an "
     "admin, not by an import, not by the agent.",
     "Treating everyone on a lease as 'the tenant' leaks a household's business to the parent who "
     "co-signed. Left as a default it gets flipped by someone trying to be helpful; as a constraint "
     "it cannot be."),
    ("R9", "Party", "reachable through", "PartyContact", "1 : N", "PartyContact.party_id → Party.party_id",
     "An inbound WhatsApp number resolves: phone → PartyContact (valid today) → Party → TenancyParty → "
     "Tenancy (active today) → Unit. Five hops, all in SQL, before any model call.",
     "If the contact is not dated, a recycled phone number hands a stranger the previous tenant's "
     "unit. This is the single most dangerous join in the system."),
    ("R10", "Tenancy", "carries many", "Obligation", "1 : N", "Obligation.tenancy_id → Tenancy.tenancy_id",
     "ארנונה, contents insurance, bank guarantee, utility accounts — each with its own validity period "
     "and its own evidence document.",
     "Attaching obligations to the unit instead of the tenancy makes them survive a move-out, and the "
     "next tenant inherits the last one's expired insurance."),
    ("R11", "Asset", "may point at", "Provider", "N : 1", "Asset.warranty_provider_id → Provider.provider_id",
     "During תקופת הבדק the contractor who installed it owes the fix, not us and not the tenant.",
     "Without it, a warranty claim becomes a phone-call archaeology exercise."),
    ("R12", "Asset", "was created by", "Document", "N : 1", "Asset.source_document_id → Document.document_id",
     "The handover protocol (פרוטוקול מסירה) is what seeds the asset register for a unit. Each asset "
     "remembers the page it came from.",
     "An asset register nobody can trace is an asset register nobody trusts."),
    ("R13", "Document", "binds to many", "any entity", "N : M, via DocumentLink", "DocumentLink(document_id, entity_type, entity_id)",
     "One lease PDF is evidence about the tenancy, the unit and both signatories. One insurance "
     "certificate is evidence about one obligation.",
     "Six nullable foreign-key columns on Document works until the seventh entity needs documents. "
     "The join table does not have that ceiling."),
    ("R14", "Building", "sets default", "warranty window", "attribute inherited", "Building.warranty_end_date, Unit may override",
     "תקופת הבדק normally runs from the building's handover, but a unit handed over late, or an asset "
     "replaced under claim, can have its own end date.",
     "One global date makes the ternary responsibility decision wrong for exactly the units most "
     "likely to be in dispute."),
    ("R15", "Project", "has many", "Building", "1 : N, OPTIONAL", "Building.project_id → Project.project_id (nullable)",
     "NEW, D2. A project groups buildings under one tender. Nullable on purpose: a standalone "
     "building leaves it empty and loses nothing. Nothing in the system requires a project to exist "
     "before it is useful.",
     "Note the one edge we have NOT solved: Space.building_id stays mandatory, so a garden genuinely "
     "shared between three cores must currently hang off one of them. If Shoham turns out to be "
     "multi-core, that is the question to reopen — and it is cheap to reopen precisely because "
     "Project already exists."),
    ("R16", "Obligation", "is typed by", "ObligationType", "N : 1", "Obligation.obligation_type_id → ObligationType.obligation_type_id",
     "NEW, D5. The kinds of obligation are rows an admin manages, not a list a developer edits. "
     "Adding 'ועד בית — Shoham model' or 'water, sub-metered' is a screen, not a release.",
     "A retired type must be deactivated, never deleted — old obligations still point at it. Deleting "
     "a type orphans every historical record that used it, which is exactly the evidence a dispute "
     "needs."),
]
sheet("RELATIONSHIPS",
      ["#", "From", "", "To", "Cardinality", "How it is written", "What it means, in plain words",
       "What breaks if we get it wrong"],
      [5, 15, 17, 20, 17, 38, 52, 52],
      [tuple(x) for x in rels])


# ---------------------------------------------------------------- FIELDS
F = []


def g(name):
    F.append(("__GROUP__", name))


def f(ent, field, typ, req, key, meaning, note=""):
    F.append((ent, field, typ, req, key, meaning, note))


g("E1 · PROJECT — optional container above Building (D2)")
f("Project", "project_id", "id", "yes", "PK", "", "")
f("Project", "name", "text", "yes", "", "What staff call it. 'Shoham — Rakefet'.", "")
f("Project", "project_code", "text", "yes", "",
  "Short code. Moved here from Building, so it has one home.",
  "The tender code lives on Project, not Building. One fact, one place.")
f("Project", "tender_ref", "text", "no", "",
  "The דירה להשכיר tender this project was won under.", "")
f("Project", "status", "enum", "yes", "", "PLANNING · ACTIVE · EXITED", "")
f("Project", "—", "—", "—", "",
  "Nothing else, on purpose.",
  "Project exists so Shoham can grow into it. We add fields when we know what it must carry, not "
  "before. See R15.")

g("E2 · BUILDING — the container")
f("Building", "building_id", "id", "yes", "PK", "Internal identifier.", "")
f("Building", "name", "text", "yes", "", "What staff call it. 'Shoham — Rakefet 12'.", "")
f("Building", "address_line", "text", "yes", "", "Street and number.", "")
f("Building", "city", "text", "yes", "", "", "")
f("Building", "project_id", "id", "no", "FK → Project",
  "The project this building belongs to, if we have modelled one.",
  "CHANGED BY D2 — replaces the old project_code text field. Nullable: a standalone building "
  "leaves it empty. Nothing downstream requires it yet.")
f("Building", "handover_date", "date", "yes", "",
  "When Dona Dom took possession from the developer.", "Starts תקופת הבדק.")
f("Building", "warranty_end_date", "date", "yes", "",
  "Default end of תקופת הבדק for everything in this building.",
  "A unit or an asset may override it — see R14.")
f("Building", "status", "enum", "yes", "", "ACTIVE · IN_CONSTRUCTION · EXITED", "")
f("Building", "unit_count", "derived", "—", "",
  "Count of units. Never stored.", "Stored counts drift the first time someone adds a unit.")

g("E3 · SPACE — every addressable place, including apartments")
f("Space", "space_id", "id", "yes", "PK", "", "")
f("Space", "building_id", "id", "yes", "FK → Building", "Which building it is in.",
  "Mandatory. See the note on R15 for the one case this does not yet cover.")
f("Space", "space_kind", "enum", "yes", "",
  "UNIT · COMMON · TECHNICAL · EXTERIOR · PARKING · STORAGE",
  "UNIT = leasable apartment. COMMON = lobby, stairwell, corridor, roof, shared shelter. "
  "TECHNICAL = pump room, electrical room, elevator machine room. EXTERIOR = garden, gate, "
  "parking deck. PARKING / STORAGE added by D3 — assignable places, which behave like the unit "
  "they are assigned to. This one field does most of the work in the responsibility rule.")
f("Space", "name", "text", "yes", "", "'Apartment 12', 'Lobby', 'Roof', 'Pump room', 'Parking P-14'.", "")
f("Space", "floor", "text", "no", "", "Kept as text — ground, basement, roof are not integers.", "")
f("Space", "access_note", "text", "no", "",
  "How a technician physically gets in: key with the super, code, tenant must be present.",
  "Cheap field, saves a wasted visit. Operational, not legal.")

g("E4 · UNIT — the leasable subset of Space")
f("Unit", "unit_id", "id", "yes", "PK, FK → Space",
  "Same key as the space it is. A unit IS a space.", "See R2.")
f("Unit", "unit_number", "text", "yes", "", "As printed on the door and written in the lease.",
  "Text, not a number: '12A' exists.")
f("Unit", "rooms", "decimal", "yes", "", "Israeli convention: 3, 3.5, 4.", "")
f("Unit", "area_sqm", "decimal", "no", "", "", "")
f("Unit", "has_mamad", "bool", "yes", "", "Does it have a ממ\"ד.",
  "Safety-relevant and it comes up in service calls about the ventilation and the door seal.")
f("Unit", "parking_space_id", "id", "no", "FK → Space",
  "The parking bay assigned to this unit.",
  "SETTLED BY D3 — bays are Space rows of kind PARKING, so they can hold a gate motor and receive "
  "service calls. Reassigning a bay is one field, and the history stays with the bay.")
f("Unit", "storage_space_id", "id", "no", "FK → Space", "The storage room assigned to this unit.",
  "SETTLED BY D3 — same treatment, kind STORAGE.")
f("Unit", "warranty_end_date", "date", "no", "",
  "Overrides the building's date when this unit was handed over separately.", "See R14.")
f("Unit", "condition_status", "enum", "yes", "",
  "READY · RENOVATION · WITHHELD",
  "SETTLED BY D6 — a plain enum, and NOT occupancy. Whether the unit is occupied is derived from "
  "tenancy dates and is never stored here.")
f("Unit", "occupancy", "derived", "—", "",
  "OCCUPIED / VACANT, from whether an active tenancy exists today.", "R6. Never a column.")

g("E5 · PARTY — a person or a company")
f("Party", "party_id", "id", "yes", "PK", "", "")
f("Party", "party_kind", "enum", "yes", "", "PERSON · COMPANY", "")
f("Party", "full_name", "text", "yes", "", "", "")
f("Party", "national_id", "text", "no", "",
  "ת.ז. for a person, ח.פ. for a company.",
  "Sensitive. Admin-only, never reachable by any agent tool, and it should be access-logged.")
f("Party", "preferred_language", "enum", "no", "",
  "he · ar · ru · fr · en",
  "Drives which language the agent opens in. Default he.")

g("E6 · PARTYCONTACT — the agent's front door")
f("PartyContact", "contact_id", "id", "yes", "PK", "", "")
f("PartyContact", "party_id", "id", "yes", "FK → Party", "", "")
f("PartyContact", "channel", "enum", "yes", "", "PHONE · EMAIL", "")
f("PartyContact", "value", "text", "yes", "", "Phone stored in E.164 — +9725…",
  "One format, always. Mixed formats break the inbound lookup silently.")
f("PartyContact", "is_primary", "bool", "yes", "", "", "")
f("PartyContact", "valid_from", "date", "yes", "", "", "")
f("PartyContact", "valid_to", "date", "no", "", "Null = still current.",
  "MUST be dated. A recycled number is how a stranger reaches someone else's apartment. See R9.")
f("PartyContact", "verified_at", "timestamp", "no", "",
  "When we last confirmed this number really is this person.", "")

g("E7 · TENANCY — one lease term on one unit")
f("Tenancy", "tenancy_id", "id", "yes", "PK", "", "")
f("Tenancy", "unit_id", "id", "yes", "FK → Unit", "", "")
f("Tenancy", "start_date", "date", "yes", "", "", "")
f("Tenancy", "end_date", "date", "yes", "", "Contractual end.",
  "Together with start_date this is the isolation window. No overlap allowed on one unit.")
f("Tenancy", "status", "enum", "yes", "",
  "DRAFT · ACTIVE · ENDED · TERMINATED_EARLY", "")
f("Tenancy", "terms_profile_id", "id", "yes", "FK → TermsProfile",
  "Which maintenance annex governs this lease.",
  "The responsibility matrix is per profile, not global. Still an open question how many exist.")
f("Tenancy", "notice_date", "date", "no", "", "When notice was given, if it was.", "")
f("Tenancy", "actual_move_out", "date", "no", "", "Reality, when it differs from end_date.", "")
f("Tenancy", "—", "—", "—", "", "NO RENT. NO DEPOSIT. NO BALANCE.",
  "Deliberate and permanent. Money lives in Priority behind read-only keys.")

g("E8 · TENANCYPARTY — who is on the lease, and in what role")
f("TenancyParty", "tenancy_id", "id", "yes", "PK part, FK", "", "")
f("TenancyParty", "party_id", "id", "yes", "PK part, FK", "", "")
f("TenancyParty", "role", "enum", "yes", "",
  "PRIMARY_TENANT · CO_TENANT · GUARANTOR · OCCUPANT",
  "GUARANTOR = ערב. On the lease, but not a resident.")
f("TenancyParty", "is_service_contact", "bool", "yes", "",
  "May this party open and discuss service calls for the unit?",
  "LOCKED BY D4 — forced FALSE whenever role = GUARANTOR, and it cannot be set true by an admin, "
  "an import or the agent. A constraint in the database, not a default in a form. See R8.")

g("E9 · OBLIGATION — the tenant-side liabilities, status only")
f("Obligation", "obligation_id", "id", "yes", "PK", "", "")
f("Obligation", "tenancy_id", "id", "yes", "FK → Tenancy", "Attached to the tenancy, not the unit.",
  "See R10.")
f("Obligation", "obligation_type_id", "id", "yes", "FK → ObligationType",
  "Which kind of obligation this is.",
  "CHANGED BY D5 — was a fixed enum, now points at an admin-managed row. See R16.")
f("Obligation", "responsible_party", "enum", "yes", "", "TENANT · OPERATOR",
  "Copied from the type when the obligation is created, then stored here. A later change to the "
  "catalogue must never rewrite an obligation that already exists.")
f("Obligation", "valid_from", "date", "no", "", "", "")
f("Obligation", "valid_to", "date", "no", "", "Policy expiry, guarantee expiry, bill period end.", "")
f("Obligation", "evidence_document_id", "id", "no", "FK → Document",
  "The certificate, bill or guarantee that proves it.", "")
f("Obligation", "status", "derived", "—", "",
  "SATISFIED · EXPIRING · EXPIRED · MISSING",
  "Computed from dates plus whether evidence exists. Not typed in by hand.")

g("E10 · OBLIGATIONTYPE — the admin-managed catalogue (D5)")
f("ObligationType", "obligation_type_id", "id", "yes", "PK", "", "")
f("ObligationType", "code", "text", "yes", "unique",
  "Stable machine key: ARNONA · CONTENTS_INSURANCE · BANK_GUARANTEE · UTILITY_ACCOUNT · "
  "HOUSE_COMMITTEE.",
  "The five seed rows. Never renamed and never reused — rules and history point at this.")
f("ObligationType", "label_he", "text", "yes", "",
  "What the admin sees: ארנונה · ביטוח תכולה · ערבות בנקאית · חשבון שירות · ועד בית.", "")
f("ObligationType", "label_en", "text", "no", "", "", "")
f("ObligationType", "default_responsible_party", "enum", "yes", "", "TENANT · OPERATOR",
  "A default for new obligations only. Copied at creation, never pushed onto existing rows.")
f("ObligationType", "requires_evidence", "bool", "yes", "",
  "Must a document exist before this can read SATISFIED?",
  "Insurance and a bank guarantee: yes. A utility account may be a yes/no we simply confirm.")
f("ObligationType", "is_active", "bool", "yes", "",
  "Retired types stay as rows and stop appearing in the 'add' list.",
  "NEVER delete a type. Old obligations still point at it, and those are the records a dispute "
  "reads. See R16.")

g("E11 · ASSET — anything that breaks, is inspected, or is handed over")
f("Asset", "asset_id", "id", "yes", "PK", "", "")
f("Asset", "space_id", "id", "yes", "FK → Space",
  "Where it is. Always exactly one space.", "The key simplification. See R3.")
f("Asset", "asset_class", "enum", "yes", "",
  "FIXTURE · SAFETY · UTILITY",
  "Your three categories, as an attribute rather than three tables. FIXTURE = AC, water heater, oven. "
  "SAFETY = extinguisher, sprinkler, smoke detector, emergency light. UTILITY = pump, elevator, "
  "boiler, gate motor, intercom.")
f("Asset", "asset_type", "enum", "yes", "",
  "The specific kind, from a controlled list.",
  "This is the field the responsibility matrix keys on. It must be a list, never free text — and "
  "unlike obligation types, this list is governed, not admin-editable. Editing it edits policy.")
f("Asset", "make_model", "text", "no", "", "", "")
f("Asset", "serial_no", "text", "no", "", "", "")
f("Asset", "installed_date", "date", "no", "", "", "")
f("Asset", "warranty_end_date", "date", "no", "",
  "When this asset leaves תקופת הבדק.",
  "Drives asset_in_warranty, which is what makes responsibility ternary instead of binary.")
f("Asset", "warranty_provider_id", "id", "no", "FK → Provider",
  "Who owes the fix while it is in warranty.", "See R11.")
f("Asset", "compliance_regime", "enum", "yes", "",
  "NONE · PERIODIC_INSPECTION",
  "Extinguishers, sprinklers, elevators and gas installations carry a legal inspection cycle. "
  "Most fixtures do not.")
f("Asset", "next_inspection_due", "date", "no", "",
  "When the next certificate is due.",
  "This single field is what makes a building screen worth opening — see the ADMIN VIEWS sheet.")
f("Asset", "last_certificate_document_id", "id", "no", "FK → Document",
  "The most recent inspection certificate.", "")
f("Asset", "source_document_id", "id", "no", "FK → Document",
  "The handover protocol that created this asset row.", "See R12.")
f("Asset", "status", "enum", "yes", "", "IN_SERVICE · FAULTY · REMOVED", "")

g("E12 · DOCUMENT — the evidence")
f("Document", "document_id", "id", "yes", "PK", "", "")
f("Document", "type_key", "enum", "yes", "",
  "lease · lease_amendment · termination_notice · arnona · insurance · id · bank_guarantee · "
  "handover_protocol · inspection_certificate",
  "The first eight are already agreed. inspection_certificate is new here, and needed by SAFETY assets.")
f("Document", "storage_uri", "text", "yes", "", "Our copy, in our object storage.", "")
f("Document", "sha256", "text", "yes", "", "Hash taken at ingest.", "Proves the copy is the file we read.")
f("Document", "drive_file_id", "text", "no", "",
  "Provenance — which Google Drive file this came from.",
  "Provenance only. Drive is a source, never the system of record, and a Drive folder name is never "
  "a binding.")
f("Document", "valid_from", "date", "no", "", "For dated documents: bill period, policy period.", "")
f("Document", "valid_to", "date", "no", "", "", "")
f("Document", "ingested_at", "timestamp", "yes", "", "", "")

g("E13 · DOCUMENTLINK — one document, several bindings")
f("DocumentLink", "document_id", "id", "yes", "PK part, FK", "", "")
f("DocumentLink", "entity_type", "enum", "yes", "",
  "PROJECT · BUILDING · SPACE · UNIT · TENANCY · PARTY · ASSET · OBLIGATION",
  "PROJECT added by D2 — a tender document belongs to the project, not to any one building.")
f("DocumentLink", "entity_id", "id", "yes", "PK part", "", "")
f("DocumentLink", "link_role", "enum", "no", "",
  "SIGNATORY · SUBJECT · EVIDENCE · SOURCE",
  "A lease links to two parties as SIGNATORY and to the unit as SUBJECT.")

g("E14 · PROVIDER — stub only, out of scope this pass")
f("Provider", "provider_id", "id", "yes", "PK", "", "")
f("Provider", "name", "text", "yes", "", "", "")
f("Provider", "provider_kind", "enum", "yes", "", "IN_HOUSE_CREW · CONTRACTOR · DEVELOPER_WARRANTY",
  "Present only to make R11 resolvable. Everything else about providers waits.")

sheet("FIELDS",
      ["Entity", "Field", "Type", "Req", "Key", "Meaning / allowed values", "Note, rule or warning"],
      [15, 28, 11, 6, 19, 46, 58],
      F)


# ---------------------------------------------------------------- ADMIN VIEWS
views = [
    ("__GROUP__", "BUILDING SCREEN — what an admin sees when they open a building"),
    ("Header", "Name, address, project, unit count, handover date",
     "Building + Project + count of Unit", "Stored / derived count",
     "Identity. Read in one second, then ignored. Project shows only when the building has one."),
    ("Header strip", "תקופת הבדק: 214 days remaining · or ENDED",
     "Building.warranty_end_date", "Derived",
     "Changes who pays for everything in the building. It belongs at the top, not on a settings page."),
    ("Tab 1", "Units — grid of all units with occupancy and lease end",
     "Unit + Tenancy (today ∈ range)", "Derived, never stored",
     "The default view. Occupancy is a colour, lease-ending-soon is a colour."),
    ("Tab 2", "Spaces & assets — the building as places, each with what is in it",
     "Space + Asset", "Stored",
     "This is the tab that only exists because of the Space idea. Lobby, stairwell, roof, pump room, "
     "parking bays, storage rooms — each expandable to its assets."),
    ("Tab 3", "Compliance — inspections due, overdue, and certificates on file",
     "Asset where compliance_regime = PERIODIC_INSPECTION", "Derived from next_inspection_due",
     "THE reason a building screen earns its place. 'Three extinguisher checks overdue, elevator "
     "permit expires in 20 days.' Nothing else in the system surfaces this."),
    ("Tab 4", "Documents — building-level papers",
     "Document via DocumentLink where entity_type = BUILDING", "Stored", ""),
    ("Tab 5", "Open service calls in this building", "ServiceCall", "Out of scope this pass",
     "Listed for completeness only."),

    ("__GROUP__", "UNIT SCREEN — the 'unit sheet', and the hardest one to get right"),
    ("Header", "Unit number, building, rooms, m², floor, ממ\"ד",
     "Unit + Space", "Stored", ""),
    ("Occupancy chip", "OCCUPIED until 31 Aug 2027 · or VACANT",
     "Tenancy where today ∈ [start, end]", "DERIVED — computed on every load",
     "This is R6 made visible. It is a chip on a screen, not a column in a table. If it is ever "
     "stored, isolation is broken and no one will notice for months."),
    ("Panel 1", "Current tenancy — who, in what role, how to reach them, in what language",
     "Tenancy + TenancyParty + Party + PartyContact", "Derived",
     "Roles shown explicitly. The guarantor appears here as a guarantor, greyed, marked 'not a "
     "service contact' — and per D4 that marking is not a toggle anyone can flip."),
    ("Panel 2", "Obligations strip — ארנונה ✓ · insurance expires in 40 days ⚠ · guarantee ✓",
     "Obligation + ObligationType + Document", "Status derived from dates",
     "Status only. No amounts, ever. An admin who needs a number opens Priority. Which obligations "
     "can appear here is set on the settings screen below, not in code."),
    ("Panel 3", "Assets in this unit, each with warranty state",
     "Asset where space_id = this unit", "Stored + derived flag",
     "'AC — in warranty until Mar 2027 — developer' is the whole responsibility answer, "
     "visible before anyone argues."),
    ("Panel 4", "Assigned parking bay and storage room, each with its own assets",
     "Space where kind = PARKING / STORAGE", "Stored",
     "New from D3. The bay is a place, so the gate motor serving it has somewhere to live and a "
     "reassignment leaves a trail."),
    ("Panel 5", "Documents, grouped by type",
     "Document via DocumentLink", "Stored",
     "Lease, amendment, IDs, insurance, handover protocol."),
    ("Panel 6", "History — every previous tenancy on this unit",
     "Tenancy where end_date < today", "Stored",
     "Nothing overwritten, so this panel is free. It is also the panel that wins a dispute."),
    ("Absent", "Rent, balance, arrears, deposit", "—", "—",
     "Deliberately not on this screen and not in this model. Priority owns money."),

    ("__GROUP__", "SETTINGS SCREEN — the lists an admin owns, and the ones they must not (D5)"),
    ("Admin-managed", "Obligation types — add, rename the label, retire",
     "ObligationType", "Stored",
     "The point of D5. Adding a ועד בית variant or sub-metered water is a form, not a release. "
     "Retiring hides it from the 'add' list and touches no existing obligation."),
    ("Guarded", "Asset types — the list the responsibility matrix keys on",
     "Asset.asset_type", "Stored, governed",
     "NOT on this screen. Editing it edits policy, and policy is versioned with effective_from so a "
     "dispute a year later still resolves. Changes go through the same route as a rule change."),
    ("Locked", "Guarantor reachability", "TenancyParty.is_service_contact", "Constraint",
     "Not a setting at all. D4 made it a database rule so there is no switch to find."),

    ("__GROUP__", "THE TEST — if the model is right, these are all one query each"),
    ("Q1", "Who lives in unit 12 today?",
     "Unit → Tenancy (today in range) → TenancyParty → Party", "—", "No stored pointer anywhere."),
    ("Q2", "This phone number just messaged us — which unit, if any?",
     "PartyContact (valid today) → Party → TenancyParty → Tenancy (active) → Unit", "—",
     "Five hops, all SQL, before any model call. The isolation join."),
    ("Q3", "What is overdue for inspection in this building?",
     "Space → Asset where next_inspection_due < today", "—", "Works because assets hang on spaces."),
    ("Q4", "Is this broken thing the tenant's problem?",
     "Asset → Space.space_kind + Asset.warranty_end_date + Tenancy.terms_profile_id", "—",
     "Three inputs, no AI, fully inspectable a year later."),
    ("Q5", "Which leases end in the next 60 days, across 1,500 units?",
     "Tenancy where end_date between today and +60", "—", "One index, whole portfolio."),
    ("Q6", "Show me every apartment with no insurance on file.",
     "Unit → Tenancy (active) → Obligation where type = CONTENTS_INSURANCE and status ≠ SATISFIED",
     "—", "The kind of question that is impossible today and trivial here."),
    ("Q7", "Which bay is assigned to unit 12, and who serviced its gate motor?",
     "Unit.parking_space_id → Space → Asset", "—",
     "New from D3, and free — a bay is a place like any other."),
]
sheet("ADMIN VIEWS",
      ["Where", "What the admin sees", "Comes from", "Stored or derived", "Why it is there"],
      [15, 52, 44, 26, 58],
      views)


# ---------------------------------------------------------------- DECISIONS
dec = [
    ("D1", "Where does a thing live?",
     "A building is a set of Spaces. Unit is the leasable kind.",
     "Every Asset carries exactly one non-null space_id. Unit.unit_id = Space.space_id, so a unit "
     "is a 1:1 extension of a space rather than a separate family of tables.",
     "One location column means no asset query ever branches on which of several nullable columns "
     "is filled, and common areas never grow into a parallel system with their own screens. "
     "Responsibility then falls out of location: UNIT is the only kind that can be the tenant's."),
    ("D2", "What sits above a Building?",
     "Project, and it is optional.",
     "Building.project_id is nullable; project_code lives on Project so the tender code has one "
     "home. PROJECT is a valid DocumentLink.entity_type. (R15)",
     "A tender can span several entrance cores under one code, and a standalone building simply "
     "leaves the link null. One edge is deliberately left open: Space.building_id is mandatory, so "
     "a garden shared between three cores hangs off one of them. If Shoham turns out to be "
     "multi-core, that is the question to reopen — and Project already being there makes it cheap."),
    ("D3", "How are parking bays and storage rooms modelled?",
     "As Space rows, with space_kind PARKING and STORAGE.",
     "An assigned bay or room follows the responsibility of the unit it is assigned to; unassigned, "
     "it is treated as COMMON. (R4)",
     "A bay is a place, not an attribute. As a Space it can hold a gate motor, take a service call "
     "in its own right, and be reassigned to another unit without losing its history."),
    ("D4", "Can a guarantor (ערב) receive service information?",
     "Never.",
     "is_service_contact is forced false whenever role = GUARANTOR — a database constraint, not a "
     "form default.",
     "There is no toggle to find, no import path that can set it, and no agent override. The "
     "settings sheet records the field as locked so that nobody goes looking for a switch."),
    ("D5", "Where does the obligation list live?",
     "In an admin-managed catalogue.",
     "ObligationType (E10) holds the types and Obligation.obligation_type_id references it (R16). "
     "A type is deactivated, never deleted, and responsible_party is copied onto the obligation at "
     "creation. Seed rows: ARNONA · CONTENTS_INSURANCE · BANK_GUARANTEE · UTILITY_ACCOUNT · "
     "HOUSE_COMMITTEE.",
     "Adding a ועד בית variant or sub-metered water is then a screen rather than a release. Copying "
     "responsible_party at creation is what stops an edit to the catalogue from rewriting an "
     "obligation that a dispute will later read."),
    ("D6", "Where does unit condition live?",
     "A plain enum on Unit.",
     "condition (RENOVATION, WITHHELD, …) is stored on Unit and is separate from occupancy, which "
     "stays derived from Tenancy dates.",
     "Condition and occupancy answer different questions and change for different reasons. Modelling "
     "condition as a kind of tenancy would force every occupancy query to handle a synthetic row."),
]
ws = sheet("DECISIONS",
           ["#", "The question", "The decision", "The rule it creates", "Why it holds"],
           [5, 38, 46, 62, 66],
           [tuple(x) for x in dec])
for row in range(2, 2 + len(dec)):
    ws.cell(row=row, column=3).font = Font(size=10, bold=True, color=GOOD)

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "dona-building-unit-model-draft.xlsx")
wb.save(OUT)
print("written:", OUT)
